"""
app/exports/excel_exporter.py — Daily bookings export to XLSX/CSV.
"""
import os
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.core.logging import get_logger
from app.models.booking import Booking
from app.models.customer import Customer
from app.models.daily_export import DailyExport

logger = get_logger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT = Font(bold=True, color="FFFFFF")

COLUMNS = [
    "booking_id", "status", "created_at", "scheduled_at",
    "customer_name", "phone", "area", "vehicle_model",
    "vehicle_category", "service_type", "price_mad", "currency",
    "address", "source_channel", "notes",
]


async def generate_daily_export(session: AsyncSession, export_date: date) -> str | None:
    """
    Generate an Excel export for all bookings created on export_date.

    Returns the file path on success, or None on error.
    """
    # Upsert daily_export record
    from sqlalchemy.exc import IntegrityError
    export_record = DailyExport(export_date=export_date, status="pending")
    session.add(export_record)
    try:
        await session.flush()
    except IntegrityError:
        await session.rollback()
        result = await session.execute(
            select(DailyExport).where(DailyExport.export_date == export_date)
        )
        export_record = result.scalar_one()
        if export_record.status == "done":
            logger.info("export_already_done", date=str(export_date))
            return export_record.file_path

    try:
        # Fetch bookings for the day (Africa/Casablanca = UTC+1, approximate with UTC range)
        import pytz
        tz = pytz.timezone("Africa/Casablanca")
        day_start = tz.localize(datetime.combine(export_date, datetime.min.time())).astimezone(UTC)
        day_end = day_start + timedelta(days=1)

        result = await session.execute(
            select(Booking, Customer)
            .join(Customer, Booking.customer_id == Customer.id)
            .where(
                Booking.created_at >= day_start,
                Booking.created_at < day_end,
            )
            .order_by(Booking.created_at.asc())
        )
        rows = result.all()

        # Build Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Bookings {export_date}"

        # Header row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws.column_dimensions[cell.column_letter].width = 18

        # Data rows
        for row_idx, (booking, customer) in enumerate(rows, start=2):
            ws.append([
                str(booking.id),
                booking.status,
                booking.created_at.strftime("%Y-%m-%d %H:%M") if booking.created_at else "",
                booking.scheduled_start.strftime("%Y-%m-%d %H:%M") if booking.scheduled_start else "",
                customer.name or "",
                customer.phone_e164,
                booking.area_name,
                booking.vehicle_model or "",
                booking.vehicle_category,
                booking.service_type,
                float(booking.price_mad),
                booking.currency,
                booking.address_text,
                "whatsapp",
                booking.notes or "",
            ])

        # Save file
        export_dir = Path(settings.export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"golav_bookings_{export_date}.xlsx"
        file_path = str(export_dir / file_name)
        wb.save(file_path)

        # Update record
        export_record.status = "done"
        export_record.file_path = file_path
        export_record.booking_count = len(rows)
        export_record.completed_at = datetime.now(UTC)

        logger.info("export_done", date=str(export_date), count=len(rows), path=file_path)
        return file_path

    except Exception as exc:
        export_record.status = "failed"
        export_record.error = str(exc)
        logger.error("export_failed", date=str(export_date), error=str(exc))
        return None
